#!/usr/bin/env python3
"""Generate excel_演習_データ.xlsx for Anything/excel.md practice."""
import sys
from datetime import date
from pathlib import Path

ROOT = Path(__file__).resolve().parent
sys.path.insert(0, str(ROOT / ".pydeps"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
wb = Workbook()

# --- 0_README ---
ws0 = wb.active
ws0.title = "0_README"
ws0["A1"] = "Excel演習データ（excel.md 対応）"
ws0["A1"].font = Font(bold=True, size=14)
ws0["A3"] = "各シートの用途と、詳細な問題文は「excel_演習_問題と解答.md」を参照してください。"
ws0["A5"] = "シート一覧"
rows_readme = [
    ("シート名", "内容"),
    ("1_売上", "SUM / SUMIF / SUMIFS / SUMPRODUCT"),
    ("2_端数_ROUND", "ROUND / ROUNDDOWN / ROUNDUP"),
    ("3_SUBTOTAL", "SUBTOTAL（フィルタ・小計行）"),
    ("4_商品マスタ", "VLOOKUP / XLOOKUP / IFERROR 用マスタ"),
    ("5_受注一覧", "VLOOKUP・LEFT・TRIM・COUNTIFS"),
    ("6_日付", "WORKDAY / NETWORKDAYS / EOMONTH / DATEDIF"),
]
for i, row in enumerate(rows_readme, start=6):
    for j, v in enumerate(row, start=1):
        ws0.cell(row=i, column=j, value=v)
ws0.column_dimensions["A"].width = 22
ws0.column_dimensions["B"].width = 50

# --- 1_売上 ---
ws1 = wb.create_sheet("1_売上")
headers1 = ["月", "担当", "カテゴリ", "金額"]
data1 = [
    [4, "田中", "備品", 12000],
    [4, "田中", "消耗品", 8000],
    [4, "佐藤", "備品", 5000],
    [5, "田中", "備品", 15000],
    [5, "佐藤", "備品", 9000],
    [5, "佐藤", "消耗品", 3000],
    [6, "田中", "消耗品", 7000],
]
for c, h in enumerate(headers1, 1):
    ws1.cell(1, c, value=h).font = Font(bold=True)
for r, row in enumerate(data1, 2):
    for c, v in enumerate(row, 1):
        ws1.cell(r, c, value=v)
ws1["F1"] = "演習用（ここに数式）"
ws1["F1"].font = Font(bold=True)
ws1["F1"].fill = PatternFill(fill_type="solid", fgColor="FFF9C4")
ws1.merge_cells("F1:G1")

# 見積用ミニ表（同一シート下部）
ws1["A12"] = "■ 見積（SUMPRODUCT 用）"
ws1["A12"].font = Font(bold=True)
for c, h in enumerate(["品目", "単価", "数量"], 1):
    ws1.cell(13, c, value=h).font = Font(bold=True)
items = [("ボルト", 12, 500), ("ナット", 8, 500), ("ワッシャ", 3, 1000)]
for r, row in enumerate(items, 14):
    for c, v in enumerate(row, 1):
        ws1.cell(r, c, value=v)

# --- 2_端数_ROUND ---
ws2 = wb.create_sheet("2_端数_ROUND")
ws2["A1"] = "税抜金額"
ws2["B1"] = "消費税率"
ws2["A2"] = 12345
ws2["B2"] = 0.1
ws2["A1"].font = ws2["B1"].font = Font(bold=True)
ws2["D1"] = "演習：税込・端数処理（D列以降に数式）"
ws2["D1"].font = Font(bold=True)
ws2["D1"].fill = PatternFill(fill_type="solid", fgColor="FFF9C4")

# --- 3_SUBTOTAL ---
ws3 = wb.create_sheet("3_SUBTOTAL")
ws3["A1"], ws3["B1"] = "部門", "金額"
ws3["A1"].font = ws3["B1"].font = Font(bold=True)
# 明細のみ（小計行なし）。フィルタ後も正しい合計になるよう SUBTOTAL(9,...) を使う演習用。
sub_data = [
    ("営業", 100000),
    ("営業", 50000),
    ("開発", 80000),
    ("開発", 20000),
    ("開発", 15000),
]
for r, (a, b) in enumerate(sub_data, 2):
    ws3.cell(r, 1, value=a)
    ws3.cell(r, 2, value=b)
ws3["A8"] = "▼フィルタをかけたあとも、見えている行だけ合計（SUBTOTAL）"
ws3["A8"].font = Font(bold=True)
ws3["D1"] = "※ B2:B6 がデータ。オートフィルタをかけ、SUBTOTAL(9,...) と SUM の違いを確認。"
ws3["D1"].font = Font(italic=True)

# --- 4_商品マスタ ---
ws4 = wb.create_sheet("4_商品マスタ")
for c, h in enumerate(["商品コード", "商品名", "標準単価"], 1):
    ws4.cell(1, c, value=h).font = Font(bold=True)
master = [
    ("P-1001", "六角ボルト M6", 15),
    ("P-1002", "六角ナット M6", 9),
    ("P-2001", "スチール板 2mm", 3200),
]
for r, row in enumerate(master, 2):
    for c, v in enumerate(row, 1):
        ws4.cell(r, c, value=v)

# --- 5_受注一覧 ---
ws5 = wb.create_sheet("5_受注一覧")
for c, h in enumerate(["入力コード", "支店", "ステータス", "備考"], 1):
    ws5.cell(1, c, value=h).font = Font(bold=True)
# 先頭2文字がカテゴリコード P- / Q- 等想定、名前に余分スペース
raw5 = [
    ("  P-1001  ", "東京", "未入金", ""),
    ("P-1002", "大阪", "入金済", ""),
    ("P-1001", "東京", "未入金", ""),
    ("Q-9001", "東京", "未入金", ""),
    ("P-2001 ", "名古屋", "入金済", ""),
]
for r, row in enumerate(raw5, 2):
    for c, v in enumerate(row, 1):
        ws5.cell(r, c, value=v)
ws5["F1"] = "演習列（クリーニング・分類・参照）"
ws5["F1"].font = Font(bold=True)
ws5["F1"].fill = PatternFill(fill_type="solid", fgColor="FFF9C4")

# --- 6_日付 ---
ws6 = wb.create_sheet("6_日付")
ws6["A1"] = "契約開始日"
ws6["B1"] = "納期（固定日）"
ws6["C1"] = "祝日リスト（WORKDAY用・必要なら参照）"
ws6["A1"].font = ws6["B1"].font = ws6["C1"].font = Font(bold=True)
ws6["A2"] = date(2026, 3, 2)
ws6["B2"] = date(2026, 4, 30)
ws6["C2"] = date(2026, 3, 20)  # 祝日例
ws6["C3"] = date(2026, 4, 29)

out_path = ROOT / "excel_演習_データ.xlsx"
wb.save(out_path)
print("Wrote", out_path)
